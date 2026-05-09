from pathlib import Path
import shutil
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook


def fill_customer_row(ws, row, seq, name, detail_row, bad_class='关联方组合'):
    """填充单个客户行"""
    ws.cell(row=row, column=1, value=seq)           # A=序号
    ws.cell(row=row, column=2, value=name)          # B=债务人名称
    ws.cell(row=row, column=3, value='销售商品')    # C=经济内容
    ws.cell(row=row, column=4, value=bad_class)     # D=期初坏账准备计提分类
    ws.cell(row=row, column=5, value=bad_class)     # E=期末坏账准备计提分类
    ws.cell(row=row, column=6, value=float(detail_row.get('本位币期初', 0)))  # F
    ws.cell(row=row, column=7, value=0)             # G
    ws.cell(row=row, column=8, value=0)             # H
    ws.cell(row=row, column=9, value=f'=F{row}+G{row}-H{row}')  # I

    ws.cell(row=row, column=10, value=float(detail_row.get('期初账龄1年以内', 0)))  # J
    ws.cell(row=row, column=11, value=float(detail_row.get('期初账龄1-2年', 0)))   # K
    ws.cell(row=row, column=12, value=float(detail_row.get('期初账龄2-3年', 0)))   # L
    ws.cell(row=row, column=13, value=float(detail_row.get('期初账龄3-4年', 0)))   # M
    ws.cell(row=row, column=14, value=float(detail_row.get('期初账龄4-5年', 0)))   # N
    ws.cell(row=row, column=15, value=float(detail_row.get('期初账龄5年+', 0)))    # O

    ws.cell(row=row, column=16, value=float(detail_row.get('本期借方', 0)))  # P
    ws.cell(row=row, column=17, value=float(detail_row.get('本期贷方', 0)))  # Q

    ws.cell(row=row, column=18, value=float(detail_row.get('本位币期末', 0)))  # R
    ws.cell(row=row, column=19, value=0)  # S
    ws.cell(row=row, column=20, value=0)  # T
    ws.cell(row=row, column=21, value=f'=R{row}+S{row}-T{row}')  # U

    ws.cell(row=row, column=22, value=float(detail_row.get('期末账龄1年以内', 0)))  # V
    ws.cell(row=row, column=23, value=float(detail_row.get('期末账龄1-2年', 0)))   # W
    ws.cell(row=row, column=24, value=float(detail_row.get('期末账龄2-3年', 0)))   # X
    ws.cell(row=row, column=25, value=float(detail_row.get('期末账龄3-4年', 0)))   # Y
    ws.cell(row=row, column=26, value=float(detail_row.get('期末账龄4-5年', 0)))   # Z
    ws.cell(row=row, column=27, value=float(detail_row.get('期末账龄5年+', 0)))    # AA

    ws.cell(row=row, column=28, value=0)  # AB 坏账准备（后续单独填）


def sum_formula(col_letter: str, rows: List[int]) -> str | int:
    """生成SUM公式"""
    if not rows:
        return 0
    return f'=SUM({col_letter}{min(rows)}:{col_letter}{max(rows)})'


def run_ar_audit(
    balance_path: str,
    ar_support_path: str,
    related_path: str,
    contract_liab_path: str,
    journal_path: str,
    template_path: str,
    output_path: str,
) -> Dict[str, Any]:
    """
    应收账款审计底稿自动填充主函数
    """
    try:
        # ============================================================
        # 1. 读取所有源数据
        # ============================================================
        df_balance = pd.read_excel(balance_path, header=0, dtype=str)
        df_balance.columns = ['科目编码', '科目名称', '期初方向', '期初余额', '本期借方', '本期贷方', '期末方向', '期末余额']
        for c in ['期初余额', '本期借方', '本期贷方', '期末余额']:
            df_balance[c] = pd.to_numeric(df_balance[c], errors='coerce').fillna(0)
        bal = df_balance.set_index('科目编码')

        ar_total = bal.loc['1122']
        bad_debt_total = bal.loc['1231']
        revenue = bal.loc['6001']

        xls_ar = pd.ExcelFile(ar_support_path)
        df_ar_bal = pd.read_excel(xls_ar, sheet_name='辅助余额表（不含重分类）')
        df_ar_detail = pd.read_excel(xls_ar, sheet_name='辅助明细表（不含重分类）')
        df_bad_calc2 = pd.read_excel(xls_ar, sheet_name='坏账计算表（含合同负债重分类）')

        df_related = pd.read_excel(related_path)
        df_cl = pd.read_excel(contract_liab_path)

        df_journal = pd.read_excel(journal_path, header=0, dtype=str)
        df_journal.columns = [
            '日期', '月', '日', '凭证号', '一级科目', '二级科目', '三级科目', '四级科目',
            '摘要', '借方金额', '贷方金额', '借方原币', '贷方原币', '币种', '科目编码',
            '唯一标识', '筛选', '辅助项', '分录号'
        ]
        for c in ['借方金额', '贷方金额', '借方原币', '贷方原币']:
            df_journal[c] = pd.to_numeric(df_journal[c], errors='coerce').fillna(0)

        # ============================================================
        # 2. 准备明细表数据：按客户分类
        # ============================================================
        df_ar_bal_data = df_ar_bal[df_ar_bal['序号'] != '合计'].copy()
        df_ar_bal_data = df_ar_bal_data.merge(df_related[['客户名称', '客户性质']], on='客户名称', how='left')

        df_ar_detail_data = df_ar_detail[df_ar_detail['序号'] != '合计'].copy()

        merge_customers = df_ar_bal_data[df_ar_bal_data['客户性质'] == '合并范围内关联方'].sort_values('客户名称')
        nonmerge_customers = df_ar_bal_data[df_ar_bal_data['客户性质'] == '非合并范围内关联方'].sort_values('客户名称')
        nonrelated_customers = df_ar_bal_data[df_ar_bal_data['客户性质'] == '非关联方'].sort_values('客户名称')

        # ============================================================
        # 3. 复制模板并加载工作簿
        # ============================================================
        output_path = str(Path(output_path))
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)

        # ============================================================
        # 3.1 填充【审定表】
        # ============================================================
        ws_audit = wb['审定表']
        ws_audit['B6'] = float(ar_total['期末余额'])
        ws_audit['B7'] = float(bad_debt_total['期末余额'])
        ws_audit['B8'] = float(ar_total['期末余额']) - float(bad_debt_total['期末余额'])

        ws_audit['F6'] = float(ar_total['期初余额'])
        ws_audit['F7'] = float(bad_debt_total['期初余额'])
        ws_audit['F8'] = float(ar_total['期初余额']) - float(bad_debt_total['期初余额'])

        # ============================================================
        # 3.2 填充【明细表】
        # ============================================================
        ws_detail = wb['明细表（按债务人分，含账龄和坏账明细）']

        merged_to_remove = []
        for mc in ws_detail.merged_cells.ranges:
            if mc.min_row >= 6 and mc.min_col <= 2 and mc.max_col >= 2:
                merged_to_remove.append(mc)
        for mc in merged_to_remove:
            ws_detail.unmerge_cells(str(mc))

        current_row = 8
        for idx, (_, cust) in enumerate(merge_customers.iterrows(), 1):
            detail = df_ar_detail_data[df_ar_detail_data['客户名称'] == cust['客户名称']]
            d = detail.iloc[0] if len(detail) > 0 else {}
            fill_customer_row(ws_detail, current_row, idx, cust['客户名称'], d, '关联方组合')
            current_row += 1
        merge_end_row = current_row - 1

        nonmerge_start_row = 28
        current_row = nonmerge_start_row
        for idx, (_, cust) in enumerate(nonmerge_customers.iterrows(), 1):
            detail = df_ar_detail_data[df_ar_detail_data['客户名称'] == cust['客户名称']]
            d = detail.iloc[0] if len(detail) > 0 else {}
            fill_customer_row(ws_detail, current_row, idx, cust['客户名称'], d, '关联方组合')
            current_row += 1
        nonmerge_end_row = current_row - 1

        nonrelated_start_row = 31
        current_row = nonrelated_start_row
        for idx, (_, cust) in enumerate(nonrelated_customers.iterrows(), 1):
            detail = df_ar_detail_data[df_ar_detail_data['客户名称'] == cust['客户名称']]
            d = detail.iloc[0] if len(detail) > 0 else {}
            fill_customer_row(ws_detail, current_row, idx, cust['客户名称'], d, '账龄组合')
            current_row += 1
        nonrelated_end_row = current_row - 1

        merge_rows = list(range(8, merge_end_row + 1))
        nonmerge_rows = list(range(nonmerge_start_row, nonmerge_end_row + 1))
        related_rows = merge_rows + nonmerge_rows
        nonrelated_rows = list(range(nonrelated_start_row, nonrelated_end_row + 1))

        # 关联方小计（行6）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T')]:
            ws_detail.cell(row=6, column=col_idx, value=sum_formula(col_letter, related_rows))
        ws_detail.cell(row=6, column=9, value='=F6+G6-H6')
        ws_detail.cell(row=6, column=21, value='=R6+S6-T6')
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=6, column=col_idx, value=sum_formula(col_letter, related_rows))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=6, column=col_idx, value=sum_formula(col_letter, related_rows))

        # 合并范围内小计（行7）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T')]:
            ws_detail.cell(row=7, column=col_idx, value=sum_formula(col_letter, merge_rows))
        ws_detail.cell(row=7, column=9, value='=F7+G7-H7')
        ws_detail.cell(row=7, column=21, value='=R7+S7-T7')
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=7, column=col_idx, value=sum_formula(col_letter, merge_rows))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=7, column=col_idx, value=sum_formula(col_letter, merge_rows))

        # 非合并范围内小计（行27）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T')]:
            ws_detail.cell(row=27, column=col_idx, value=sum_formula(col_letter, nonmerge_rows))
        ws_detail.cell(row=27, column=9, value='=F27+G27-H27')
        ws_detail.cell(row=27, column=21, value='=R27+S27-T27')
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=27, column=col_idx, value=sum_formula(col_letter, nonmerge_rows))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=27, column=col_idx, value=sum_formula(col_letter, nonmerge_rows))

        # 非关联方小计（行30）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T')]:
            ws_detail.cell(row=30, column=col_idx, value=sum_formula(col_letter, nonrelated_rows))
        ws_detail.cell(row=30, column=9, value='=F30+G30-H30')
        ws_detail.cell(row=30, column=21, value='=R30+S30-T30')
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=30, column=col_idx, value=sum_formula(col_letter, nonrelated_rows))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=30, column=col_idx, value=sum_formula(col_letter, nonrelated_rows))

        # 账面余额总计（行143）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (9, 'I'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T'), (21, 'U')]:
            ws_detail.cell(row=143, column=col_idx, value=sum_formula(col_letter, [6, 30]))
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=143, column=col_idx, value=sum_formula(col_letter, [6, 30]))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=143, column=col_idx, value=sum_formula(col_letter, [6, 30]))

        # 坏账准备（行144）
        ws_detail.cell(row=144, column=18, value=float(df_bad_calc2['应计提坏账'].sum()))
        ws_detail.cell(row=144, column=21, value='=R144')

        # 账面价值总计（行145）
        for col_idx, col_letter in [(6, 'F'), (7, 'G'), (8, 'H'), (9, 'I'), (16, 'P'), (17, 'Q'), (18, 'R'), (19, 'S'), (20, 'T'), (21, 'U')]:
            ws_detail.cell(row=145, column=col_idx, value=sum_formula(col_letter, [143, 144]))
        for col_idx, col_letter in [(10, 'J'), (11, 'K'), (12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:
            ws_detail.cell(row=145, column=col_idx, value=sum_formula(col_letter, [143, 144]))
        for col_idx, col_letter in [(22, 'V'), (23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z'), (27, 'AA')]:
            ws_detail.cell(row=145, column=col_idx, value=sum_formula(col_letter, [143, 144]))

        # 非关联方坏账准备（AB列）
        for r in nonrelated_rows:
            name = ws_detail.cell(row=r, column=2).value
            if name:
                bad_row = df_bad_calc2[df_bad_calc2['客户名称'] == name]
                if len(bad_row) > 0:
                    ws_detail.cell(row=r, column=28, value=float(bad_row.iloc[0]['应计提坏账']))

        # 审计说明区域：关联方占比
        ws_detail.cell(row=150, column=3, value='=IF(U143=0,0,U6/U143)')

        # 重分类调整
        cl_total_end = float(df_cl[df_cl['序号'] != '合计']['本位币期末'].sum())
        ws_detail.cell(row=153, column=4, value=cl_total_end)
        ws_detail.cell(row=155, column=4, value=cl_total_end)

        # 前五名排名
        top5 = df_ar_bal_data.nlargest(5, '本位币期末')
        for i, (_, cust) in enumerate(top5.iterrows()):
            rank_row = 163 + i
            ws_detail.cell(row=rank_row, column=3, value=float(cust['本位币期末']))
            ws_detail.cell(row=rank_row, column=4, value=f'=IF(U143=0,0,C{rank_row}/U143)')
        ws_detail.cell(row=168, column=3, value='=SUM(C163:C167)')
        ws_detail.cell(row=168, column=4, value='=IF(U143=0,0,C168/U143)')

        # ============================================================
        # 3.3 填充【坏账测算表】
        # ============================================================
        ws_bad = wb['坏账测算表']

        bad_calc_data = df_bad_calc2.copy()
        num_cols = ['1年以内', '1-2年', '2-3年', '3-4年', '4-5年', '5年+', '重分类后余额', '应计提坏账']
        for c in num_cols:
            if c in bad_calc_data.columns:
                bad_calc_data[c] = pd.to_numeric(bad_calc_data[c], errors='coerce').fillna(0)

        if '合计' in bad_calc_data['客户名称'].values or '合计' in bad_calc_data['序号'].values:
            bad_calc_data = bad_calc_data[
                ~bad_calc_data['客户名称'].isin(['合计']) & ~bad_calc_data['序号'].isin(['合计'])
            ]

        age_1y = float(bad_calc_data['1年以内'].sum())
        age_1_2y = float(bad_calc_data['1-2年'].sum())
        age_2_3y = float(bad_calc_data['2-3年'].sum())
        age_3_4y = float(bad_calc_data['3-4年'].sum())
        age_4_5y = float(bad_calc_data['4-5年'].sum())
        age_5y_plus = float(bad_calc_data['5年+'].sum())
        total_bad_debt = float(bad_calc_data['应计提坏账'].sum())

        ws_bad.cell(row=25, column=8, value=age_1y)
        ws_bad.cell(row=26, column=8, value=age_1_2y)
        ws_bad.cell(row=27, column=8, value=age_2_3y)
        ws_bad.cell(row=28, column=8, value=age_3_4y)
        ws_bad.cell(row=29, column=8, value=age_4_5y)
        ws_bad.cell(row=30, column=8, value=age_5y_plus)

        ws_bad.cell(row=25, column=12, value='=H25*K25')
        ws_bad.cell(row=26, column=12, value='=H26*K26')
        ws_bad.cell(row=27, column=12, value='=H27*K27')
        ws_bad.cell(row=28, column=12, value='=H28*K28')
        ws_bad.cell(row=29, column=12, value='=H29*K29')
        ws_bad.cell(row=30, column=12, value='=H30*K30')

        ws_bad.cell(row=31, column=8, value='=SUM(H25:H30)')
        ws_bad.cell(row=31, column=12, value='=SUM(L25:L30)')

        related_end_bal = float(merge_customers['本位币期末'].sum()) + float(nonmerge_customers['本位币期末'].sum())
        ws_bad.cell(row=38, column=8, value=related_end_bal)
        ws_bad.cell(row=38, column=12, value=0)

        ws_bad.cell(row=76, column=3, value=total_bad_debt)
        ws_bad.cell(row=76, column=4, value=total_bad_debt)
        ws_bad.cell(row=77, column=3, value=0)
        ws_bad.cell(row=77, column=4, value=0)
        ws_bad.cell(row=79, column=3, value='=SUM(C76:C78)')
        ws_bad.cell(row=79, column=4, value='=SUM(D76:D78)')
        ws_bad.cell(row=81, column=3, value='=C79+C80')
        ws_bad.cell(row=81, column=4, value='=D79+D80')
        ws_bad.cell(row=81, column=5, value='=D81-C81')

        # ============================================================
        # 3.4 填充【分析表】
        # ============================================================
        ws_analysis = wb['分析表']

        ar_begin = float(ar_total['期初余额'])
        ar_end = float(ar_total['期末余额'])
        rev = float(revenue['本期贷方'])

        ws_analysis['B6'] = ar_begin
        ws_analysis['B7'] = ar_end
        ws_analysis['B8'] = rev
        ws_analysis['B9'] = '=IF(B8=0,0,B7/B8)'
        ws_analysis['C9'] = '0.00%'
        ws_analysis['B10'] = '=IF((B6+B7)/2=0,0,B8/((B6+B7)/2))'
        ws_analysis['C10'] = '0.00'
        ws_analysis['B11'] = '=IF(B10=0,0,365/B10)'
        ws_analysis['C11'] = '0.00'

        ws_analysis['B17'] = 0.13
        ws_analysis['B18'] = '=B8*(1+B17)'
        ws_analysis['B19'] = float(ar_total['本期借方'])
        ws_analysis['B20'] = '=B18-B19'

        # ============================================================
        # 3.5 填充【函证结果汇总表】
        # ============================================================
        ws_confirm = wb['函证结果汇总表']

        merged_to_remove_c = []
        for mc in ws_confirm.merged_cells.ranges:
            if mc.min_row >= 5:
                merged_to_remove_c.append(mc)
        for mc in merged_to_remove_c:
            ws_confirm.unmerge_cells(str(mc))

        all_customers = pd.concat([merge_customers, nonmerge_customers, nonrelated_customers])
        current_row_c = 5
        for idx, (_, cust) in enumerate(all_customers.iterrows(), 1):
            ws_confirm.cell(row=current_row_c, column=1, value=f'询证函-{idx:03d}')
            ws_confirm.cell(row=current_row_c, column=2, value=cust['客户名称'])
            ws_confirm.cell(row=current_row_c, column=3, value='选取')
            ws_confirm.cell(row=current_row_c, column=4, value=float(cust['本位币期末']))
            ws_confirm.cell(row=current_row_c, column=5, value=float(cust['本位币期末']))
            ws_confirm.cell(row=current_row_c, column=6, value=0)
            ws_confirm.cell(row=current_row_c, column=7, value=0)
            current_row_c += 1

        confirm_data_start = 5
        confirm_data_end = current_row_c - 1
        total_row = current_row_c
        ws_confirm.cell(row=total_row, column=1, value='合    计')
        ws_confirm.cell(row=total_row, column=4, value=f'=SUM(D{confirm_data_start}:D{confirm_data_end})')
        ws_confirm.cell(row=total_row, column=5, value=f'=SUM(E{confirm_data_start}:E{confirm_data_end})')
        ws_confirm.cell(row=total_row, column=6, value=f'=SUM(F{confirm_data_start}:F{confirm_data_end})')
        ws_confirm.cell(row=total_row, column=7, value=f'=SUM(G{confirm_data_start}:G{confirm_data_end})')

        ratio_row = total_row + 1
        ws_confirm.cell(row=ratio_row, column=1, value='回函确认金额占应收账款余额比例：')
        ws_confirm.cell(row=ratio_row, column=4, value=f'=IF(D{total_row}=0,0,E{total_row}/D{total_row})')
        ws_confirm.cell(row=ratio_row, column=7, value='替代测试确认金额占应收账款余额比例：')
        ws_confirm.cell(row=ratio_row, column=8, value=f'=IF(D{total_row}=0,0,F{total_row}/D{total_row})')

        # ============================================================
        # 4. 保存文件
        # ============================================================
        wb.save(output_path)

        return {
            "status": "success",
            "message": "底稿填充完成",
            "output_file": output_path,
            "summary": {
                "ar_total": float(ar_total['期末余额']),
                "bad_debt_total": float(bad_debt_total['期末余额']),
                "book_value": float(ar_total['期末余额']) - float(bad_debt_total['期末余额']),
                "merge_customer_count": int(len(merge_customers)),
                "nonmerge_customer_count": int(len(nonmerge_customers)),
                "nonrelated_customer_count": int(len(nonrelated_customers)),
                "confirm_customer_count": int(len(all_customers)),
                "revenue": rev,
                "bad_debt_calc_total": total_bad_debt,
                "related_end_bal": related_end_bal,
            }
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }